from django.shortcuts import render
from contact.models import Agent_Master,File_Data
from rest_framework.views import APIView
from rest_framework.response import Response
from contact.serializer import AgentMasterSerializer,AgentMasterListSerializer,ExcelrecordSerializer,ExcelrecordListSerializer
from django.http import Http404
from rest_framework.parsers import MultiPartParser,FileUploadParser
import openpyxl
from django.http import HttpResponse
# from contact.classes.celery_threat import set_reminder_task
# from openpyxl import Workbook
# from openpyxl.writer.excel import save_virtual_workbook
import os
import uuid
from django.conf import settings 
from datetime import datetime
# Create your views here.

class Agent_master_View(APIView):

    def get(self, request,formate=None):
        try:
            snippets = Agent_Master.objects.filter(is_status=True)
            serializer = AgentMasterListSerializer(snippets, many=True)
            return Response({"result": 1,"data":[serializer.data], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})
    def post(self, request, format=None):

        try:
            data=request.data
            serializer=AgentMasterSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                # id=uuid.UUID()
                path = os.path.join(settings.MEDIA_ROOT, str(serializer.data['agent_id']))
                os.mkdir(path)
                return Response({"result": 1,"data": serializer.data, "error": ""})
            return Response({"result": 0,"data": [], "error": str(serializer.errors)})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})

class Agent_masterupdate_View(APIView):

    def get_object(self, pk):
        try:
            return AgentMasterSerializer.objects.get(pk=pk)
        except AgentMasterSerializer.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        try:
            querySet = self.get_object(pk)
            serializer = AgentMasterListSerializer(querySet)
            return Response({"result": 1,"data":serializer.data, "error": ""})
        except Exception as e:
            
            return Response({"result": 0,"data": [], "error": str(e)})

    def put(self, request, pk, format=None):
        querySet = self.get_object(pk)
        serializer = AgentMasterSerializer(querySet, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"result": 1, "data":serializer.data, "error": ""})
        return Response({"result": 0, "data": [], "error": serializer.errors})

    def patch(self, request, pk, format=None):
        querySet = self.get_object(pk)
        old_priority = querySet.priority
        serializer = AgentMasterSerializer(querySet, data=request.data, partial=True) # set partial=True to update a data partially
        if serializer.is_valid():
            serializer.save()
            
            return Response({"result": 1, "data":serializer.data, "error": ""})
        return Response({"result": 0, "data": [], "error": serializer.errors})

    def delete(self, request, pk, format=None):
        try:
            querySet = self.get_object(pk)
            data={"is_status":False}
            serializer = AgentMasterSerializer(querySet, data=data, partial=True) # set partial=True to update a data partially
            if serializer.is_valid():
                serializer.save()
            return Response({"result": 1,"data":["Sucess"], "error": ""})
        except Exception as e:
           
            return Response({"result": 0,"data": [], "error": str(e)})

class Agent_Gerate_New_Id(APIView):

    def get(self, request,formate=None):
        try:
            snippets = Agent_Master.objects.filter(is_status=True).count()+1
            add_zero=['00000','0000','000','00','0']
            if len(str(snippets))<6:
                snippets=add_zero[len(str(snippets))-1]+str(snippets)
            return Response({"result": 1,"data": [str(snippets)], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})
    


class Agent_File_Upload_View(APIView):

    def get(self, request,formate=None):
        try:
            snippets = File_Data.objects.filter(is_status=True)
            serializer = ExcelrecordListSerializer(snippets, many=True)
            return Response({"result": 1,"data": [serializer.data], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})
    
    parser_classes = (MultiPartParser,FileUploadParser)

    
    def post(self, request, format=None):
        try:    
            excel_file=request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            print(request.data['agent_id'])
            excel_data = list()
            msg_error=''
            for sh,li in enumerate(wb.sheetnames):
                worksheet = wb[li]
                if sh==0:
                    headelis=worksheet[1]
                    print(worksheet[1][0])
                for j,row in enumerate(worksheet.iter_rows()):
                    if sh==0 and j==0:
                        continue
                    row_data = list()
                    dic={}
                    for i,cell in enumerate(row):
                        dic.update({str(headelis[i].value):str(cell.value)})
                    if File_Data.objects.filter(phone=dic['Phone']).exists():
                        msg_error='Some phone numbers already exists'
                        continue
                    excel_data.append({'name':dic['Name'],'phone':dic['Phone'],'agent_id':request.data['agent_id'],'xl_data':dic})
            serilizer=ExcelrecordSerializer(data=excel_data,many=True)
            if serilizer.is_valid():
                serilizer.save()
                return Response({"result": 1,"data": [serilizer.data], "error": msg_error})
            return Response({"result": 0,"data": [], "error": str(serilizer.errors)})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})

class Agent_File_dataList_View(APIView):

    def get(self, request,agent_id,formate=None):
        try:
            snippets = File_Data.objects.filter(is_status=True,agent_id=agent_id)
            serializer = ExcelrecordListSerializer(snippets, many=True)
            return Response({"result": 1,"data": [serializer.data], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})

class Campaign_dataList_View(APIView):

    def get(self, request,formate=None):
        try:
            lis =File_Data.objects.filter(is_status=True).exclude(campaign="").values_list('campaign', flat=True).distinct()
            return Response({"result": 1,"data": [lis], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})

class Dashboard_List(APIView):

    def get(self, request,formate=None):
        try:
            dic=json.loads(request.GET.get('data'))
            datetime_object = datetime.strptime(dic['date'], '%Y-%m-%d')
            snippets = Agent_Master.objects.filter(is_status=True)
            dash_list=[]
            for li in snippets:
                query_set = File_Data.objects.filter(is_status=True,agent_id=li.agent_id,added_at__year=datetime_object.year,added_at__month=datetime_object.month,added_at__day=datetime_object.day).count()
                campaign = File_Data.objects.filter(is_status=True,agent_id=li.agent_id,campaign="",is_processed=False,added_at__year=datetime_object.year,added_at__month=datetime_object.month,added_at__day=datetime_object.day).count()
                non_eligibility_2=File_Data.objects.filter(is_status=True,agent_id=li.agent_id,campaign="",is_processed=True,added_at__year=datetime_object.year,added_at__month=datetime_object.month,added_at__day=datetime_object.day).count()
                process=query_set-campaign
                process=process-non_eligibility_2
                percentage=0
                if query_set!=0:
                    percentage=round((process/query_set)*100)
                lis={
                    'agent_id':li.name+"("+str(li.agent_id)+")",
                    'total':query_set,
                    'not_in_use':campaign,
                    'processed':query_set-campaign,
                    'eligibility_non':non_eligibility_2,
                    'eligibility':process,
                    'percentage':percentage

                }
                dash_list.append(lis)
            
            return Response({"result": 1,"data": [dash_list], "error": ""})
        except Exception as e:
            return Response({"result": 0,"data": [], "error": str(e)})

import json

class Export_Contact_data(APIView):

    def get(self, request,formate=None):
        try:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "packing_list_d..xlsx"
            response['Content-Disposition'] = 'attachment; filename="'+ file_name +'"'
            dic=json.loads(request.GET.get('date'))
            datetime_object = datetime.strptime(dic['campaign_date'], '%Y-%m-%d')
            snippets = File_Data.objects.filter(is_status=True,added_at__year=datetime_object.year,added_at__month=datetime_object.month,added_at__day=datetime_object.day,campaign=dic['campaign'])
            # snippets = File_Data.objects.filter(campaign=agent_id)
            serializer = ExcelrecordListSerializer(snippets, many=True)
            wb = openpyxl.Workbook()
            sheet = wb.active

            sheet.row_dimensions[1].font = openpyxl.styles.Font(bold = True)
            headings=["name",'phone','agent_id','xl_data','campaign','is_status']
            for colno, heading in enumerate(headings, start = 1):
                sheet.cell(row = 1, column = colno).value = heading
            
            for rowno, row in enumerate(serializer.data, start = 2):
                
                for colno, cell_value in enumerate(row, start = 1):
                    sheet.cell(row = rowno, column = colno).value = str(row[cell_value])
            wb.save(response)
            return response
            # return Response({"result": 1,"data": [serializer.data], "error": ""})
        except Exception as e:
                        return Response({"result": 0,"data": [], "error": str(e)})

class Export_Contact_Viewdata(APIView):

    def get(self, request,formate=None):
        try:
            dic=json.loads(request.GET.get('date'))
            datetime_object = datetime.strptime(dic['campaign_date'], '%Y-%m-%d')
            query_set = Agent_Master.objects.filter(is_status=True)
            dash_list=[]
            for li in query_set:
                snippets = File_Data.objects.filter(added_at__year=datetime_object.year,added_at__month=datetime_object.month,added_at__day=datetime_object.day,agent_id=li.agent_id,campaign=dic['campaign'])
                lis={
                    'agent_id':li.name+"("+str(li.agent_id)+")",
                    'total':len(snippets),
                }
                dash_list.append(lis)
            return Response({"result": 1,"data": [dash_list], "error": ""})
        except Exception as e:
                        return Response({"result": 0,"data": [], "error": str(e)})